"""
Testing phase, to run from manhour.py along with leave_parser.py after finalised
1   categorise leave entries by week, excluding 'Cancel' entries, storing the data in a new dictionary
    format: {emp1: [wk1, wk2, wk3, wk4, ...], ...}
    (default wk values should be 0, wk5 should be None if it does not exist)
    (emp should have undergone strip() and lower() to prepare for matching, remove brackets?)
    (potentially, parts of the name can be separated with an underscore and the parts sorted alphabetically)
    (matching should have a tolerance that provides accuracy with minimal failure)
    (sort into 4 or 5 weeks -- use weekly_dates.py for realistic scenarios)
    (3 possibilities: 0.5 or 1 day (handled with start date), > 2 days (handled with range inclusive, edited by name)
    (provide for PH - it's in a sep column)
2   generate report by making edits to the template (for the correct no. of weeks)
    (only start editing after creating a copy of the template)
    (fill in dates for every name)
    (fill report in a linear fashion, so that the first employee is the one first listed in the template)
    (consider generating a map of employees first, so that the coordinates are readily available)
    (report should be generated on the fly (per employee) to avoid fatal errors that cripple the program)
    (ensure correct number of lines per emp, based on 4/5 wks)
    (remove from dict after edit)
3   export to json leave_summary, with totals

def export_leave_summary():
    with open('leave_summary.json', 'w') as f:
        dump(leave_summary, f)
3   to upload sample leave_report and manhour_report, rename manhour.py to main.py
4   reorg functions into right files and tidy up code - pass args through a dict for start, end dates
5   provide for hired date - approved column
6   pending leave - treated as consumed
7   alternative: parse start and end dates from leave report
8   check handling for shift work
"""

from json import load
import datetime as dt
import openpyxl as xl


def load_holidays(year):  # the right year
    holidays = []
    with open(f'ph_{year}.json') as file:
        holidays_str = load(file)
        for date in holidays_str:
            holidays.append(dt.datetime.strptime(date, '%Y/%m/%d'))
    return holidays


def remove_cancel(emp_summary):
    for emp in emp_summary.keys():
        for entry in list(emp_summary[emp].keys()):
            if emp_summary[emp][entry][2] == 'Cancel':
                del emp_summary[emp][entry]
                continue
            else:
                continue


def entry_dates(start_dt, end_dt):
    dates = []
    day = start_dt
    while True:
        dates.append(day)
        if day == end_dt:
            break
        day += dt.timedelta(days=1)
    return dates


def entry_duration():  # only working days, and excluding PH
    pass


def entry_match_week():
    pass


def transform_emp_names():  # upper, strip(), in keys
    pass


def load_template(weeks):
    template = xl.load_workbook(f'mh_template_{weeks}wk.xlsx')
    report_wb = template
    return report_wb


def write_title_date(sheet, first_day, last_day):  # change dates in other sheets too (prep template)
    cell = sheet['A1']
    start = dt.datetime.strftime(first_day, '%d/%m/%Y')
    end = dt.datetime.strftime(last_day, '%d/%m/%Y')
    cell.value = f'MANHOUR REPORT FOR CSE-ITS PTE LTD - JUN ({start} ~ {end})'


def report_focus_next_line(sheet, current_row):
    sheet.cell(row=current_row + 1, column=1)


def report_verify_line_count(weeks, sheet, row, column):
    first_cell = sheet.cell(row=row, column=column)
    count = 0
    for row in sheet[f'A{row+1}:A{row+4}']:
        for cell in row:
            if first_cell.value == cell.value:
                count += 1
    if weeks - 1 == count:
        return 1
    else:
        return 0


def report_simple_match(sheet, rows, name):
    for row in rows:
        cell_text = sheet[f'A{row}'].value.strip().upper()
        if cell_text == name.upper().strip():
            return 1, name, row
    return 0, cell_text, None


def report_match_failsafe(sheet, rows, name):
    name_parts = name.upper().split()
    name_parts_len = len(name_parts)
    exclude = ['SYED']
    for row in rows:
        matches = 0
        candidate_parts = []
        cell_text = sheet[f'A{row}'].value
        cell_text_parts = cell_text.upper().split()
        for part in cell_text_parts:
            candidate_parts.append(''.join(filter(str.isalpha, part)))
        candidate_parts_len = len(candidate_parts)
        if name_parts_len > candidate_parts_len:
            required = candidate_parts_len
        elif name_parts_len < candidate_parts_len:
            required = name_parts_len
        else:
            required = name_parts_len
        if required > 3:
            required = 3
        for part in name_parts:
            if part in candidate_parts:
                if part not in exclude:
                    matches += 1
                if matches == required:
                    return 1, name, row
    return 0, cell_text, None


def fill_holidays(sheet, holidays, row, weeks_dates):  # remove added?
    for holiday in holidays:
        for week in weeks_dates:
            if holiday in weeks_dates[week]:
                try:
                    sheet[f'F{row + week - 1}'].value += 8.3
                except TypeError:
                    sheet[f'F{row + week - 1}'].value = 8.3

                print(row, sheet[f'F{row + week - 1}'].value)


def report_add_entry(sheet, first_row, weeks_dates, dates, entry_type):  # add not replace, remove successful
    if entry_type == 'HOSPITALISATION' or entry_type == 'SICK LEAVE':
        type_column = 'L'
    elif entry_type == 'BIRTHDAY LEAVE' or entry_type == 'OIL':  # check OIL
        type_column = 'M'
    elif entry_type == 'NATIONAL SERVICE LEAVE':
        type_column = 'J'
    else:
        type_column = 'K'
    for date in dates:
        for week in weeks_dates:
            if date in weeks_dates[week]:
                try:
                    sheet[f'{type_column}{first_row + week - 1}'].value += 8.3
                except TypeError:
                    sheet[f'{type_column}{first_row + week - 1}'].value = 8.3


def entry_remove_successful():  # only if successful, based on type (Others: OIL, BIRTHDAY LEAVE;
    # HOSPITALISATION LEAVE AND SICK LEAVE; NATIONAL SERVICE LEAVE)
    pass


def append_leave_report():
    pass


def save_report():
    pass


def entry_print_remaining():  # for skipped or failed
    pass
