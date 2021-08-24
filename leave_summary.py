import xlrd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from json import dump


def read_xls_to_xlsx(source, sheet_name):
    xls = xlrd.open_workbook(source)
    xlsx = Workbook()

    xls_sheet = xls.sheet_by_name(sheet_name)
    xlsx_sheet = xlsx.active
    xlsx_sheet.title = sheet_name

    for row in range(xls_sheet.nrows):
        for column in range(xls_sheet.ncols):
            xlsx_sheet.cell(row=row+1, column=column+1).value = xls_sheet.cell_value(row, column)

    return xlsx


def load_file():
    file_path = input('Drag and drop file, then press enter: ')
    try:
        return load_workbook(file_path, read_only=True)
    except openpyxl.utils.exceptions.InvalidFileException:
        return read_xls_to_xlsx('test.xls', 'Sheet1')


def generate_leave_summary():
    global leave_summary
    leave_summary = {}
    worksheet_leave = file['Sheet1']
    # search cells for names (nested loop to search every cell)
    for row in worksheet_leave.rows:
        for cell in row:
            if str(cell.value).strip() == 'Name :':
                # search for subtotals in the area of 24 rows below each name (header is in the next column)
                for search_row in range(cell.row + 1, cell.row + 25):
                    search_cell = worksheet_leave.cell(row=search_row, column=(cell.column + 1))
                    if str(search_cell.value) == 'Subtotal':
                        # output matching names and subtotals to dictionary and returning to the second loop
                        leave_summary[worksheet_leave.cell(row=cell.row, column=(cell.column + 1)).value] = []
                        leave_summary[worksheet_leave.cell(row=cell.row, column=(cell.column + 1)).value].append(
                            worksheet_leave.cell(row=search_row, column=(cell.column + 2)).value)
                        break

    return leave_summary


def export_leave_summary():
    with open('leave_summary.json', 'w') as f:
        dump(leave_summary, f)


def print_leave_summary():
    for person in sorted(leave_summary.keys()):
        print(person, '-->', leave_summary[person])


file = load_file()
leave_summary = generate_leave_summary()
export_leave_summary()
print_leave_summary()
