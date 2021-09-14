import xlrd
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import re


def read_xls_to_xlsx(source, sheet):
    xls = xlrd.open_workbook(source)
    xlsx = Workbook()

    xls_sheet = xls.sheet_by_index(sheet)
    xls_sheet_name = xls.sheet_names()[0]
    xlsx_sheet = xlsx.active
    xlsx_sheet.title = xls_sheet_name

    for row in range(xls_sheet.nrows):
        for column in range(xls_sheet.ncols):
            xlsx_sheet.cell(row=row + 1, column=column + 1).value = xls_sheet.cell_value(row, column)

    return xlsx, xls_sheet_name


def load_file():
    file_path = input('Drag and drop leave report for staff, then press enter: ')
    try:
        return load_workbook(file_path), None
    except openpyxl.utils.exceptions.InvalidFileException:
        return read_xls_to_xlsx(file_path, 0)


def load_file_boss():
    file_path = input('Drag and drop leave report for bosses, then press enter: ')
    try:
        return load_workbook(file_path), None
    except openpyxl.utils.exceptions.InvalidFileException:
        return read_xls_to_xlsx(file_path, 0)


def parse_for_employees(file_sheet):
    data = {}
    for row in file_sheet['D10':'D3000']:  # set max with dimensions
        for cell in row:
            if cell.value == 'Name : ':
                data[file_sheet['E' + str(cell.row)].value] = cell.coordinate
    return data


def parse_for_leave(emp, emp_summary, file_sheet):
    first_entry_start = file_sheet['E' + str(file_sheet[emp_summary[emp]].row + 4)]
    if re.match('^[0-9]{2}/[0-9]{2}/[0-9]{4}$', first_entry_start.value):
        pass
    else:
        first_entry_start = file_sheet['E' + str(first_entry_start.row + 1)]
    print(emp, first_entry_start.coordinate, first_entry_start.value)
    first_entry_end = file_sheet['F' + str(first_entry_start.row)]
    first_entry_days = file_sheet['G' + str(first_entry_start.row)]
    first_entry_approval = file_sheet['I' + str(first_entry_start.row)]
    first_entry_type = file_sheet['A' + str(first_entry_start.row)]
    emp_summary[emp] = {first_entry_start.value: [first_entry_end.value, first_entry_days.value,
                                                  first_entry_approval.value, first_entry_type.value]}
    cell_in_focus = first_entry_start
    cell_attempt = cell_in_focus
    counter = 0
    while True:
        counter += 1
        cell_in_focus = cell_attempt
        for adjustment in range(1, 4):
            cell_attempt = file_sheet.cell(row=cell_in_focus.row + adjustment, column=cell_in_focus.column)
            if re.match('^[0-9]{2}/[0-9]{2}/[0-9]{4}$', str(cell_attempt.value)):
                cell_attempt_end = file_sheet['F' + str(cell_attempt.row)]
                cell_attempt_days = file_sheet['G' + str(cell_attempt.row)]
                cell_attempt_approval = file_sheet['I' + str(cell_attempt.row)]
                cell_attempt_type = file_sheet['A' + str(cell_attempt.row)]
                if cell_attempt.value not in emp_summary[emp]:
                    emp_summary[emp][cell_attempt.value] = [cell_attempt_end.value, cell_attempt_days.value,
                                                            cell_attempt_approval.value, cell_attempt_type.value]
                else:
                    emp_summary[emp][cell_attempt.value + str(counter)] = [cell_attempt_end.value, cell_attempt_days.value,
                                                                      cell_attempt_approval.value,
                                                                      cell_attempt_type.value]
                break
            if cell_attempt.value != 'Subtotal':
                continue
            break
        if cell_attempt.value == 'Subtotal':
            break
